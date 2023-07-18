import openpyxl
#Pegando a planilha que irá ser analizada
wb = openpyxl.load_workbook('v0004.xlsx')
sheet_names = wb.sheetnames
ws = wb[sheet_names[0]]
#Definindo as células que serão analizadas
col1_values = [cell.value for cell in ws['A']]
col2_values = [cell.value for cell in ws['B']]
# repetidos = []

for index, value1 in enumerate(col1_values):
  for value2 in col2_values:
    # print(value1, value2, value1 == value2)
    if value1 == value2:
      # repetidos.append(col1_values[index])
      col1_values[index] = None
      # print(index)

# print('Repetidos: \n')
# for i in repetidos:
#   print(i)
# Usando uma função lambda para retirar os valores nulos da exibição
# final = list(filter(lambda x: x is not None, col1_values))

print('\ngit initLista final sem repetidos: \n')
cont = 0
for index, value in enumerate(col1_values):
  if value is not None:
    cont+=1
    print(cont, value, 'Linha: ', 'A'+str(index+3))